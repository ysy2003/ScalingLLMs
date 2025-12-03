"""
Evaluate efficiency metrics for Gemini API results from gemini_results1.xlsx
"""
import sys
from pathlib import Path

# Add parent directory to path to import metrics
sys.path.insert(0, str(Path(__file__).parent.parent))

from metrics.efficiency import EfficiencyScores, compute_efficiency_api

try:
    import pandas as pd
    USE_PANDAS = True
except (ImportError, ValueError):
    USE_PANDAS = False

try:
    from openpyxl import load_workbook
    USE_OPENPYXL = True
except ImportError:
    USE_OPENPYXL = False
    if not USE_PANDAS:
        print("Error: Need either pandas or openpyxl to read Excel files")
        exit(1)

# Gemini pricing (per 1M tokens)
INPUT_PRICE_PER_1M = 0.30  # $0.30 per 1M tokens for input
OUTPUT_PRICE_PER_1M = 2.50  # $2.50 per 1M tokens for output

# Convert to per 1k tokens (divide by 1000)
INPUT_PRICE_PER_1K = INPUT_PRICE_PER_1M / 1000.0  # $0.0003 per 1k tokens
OUTPUT_PRICE_PER_1K = OUTPUT_PRICE_PER_1M / 1000.0  # $0.0025 per 1k tokens


def calculate_gemini_efficiency(data) -> EfficiencyScores:
    """
    Calculate efficiency scores for Gemini API results.
    
    Args:
        data: DataFrame or list of dicts with columns: prompt_token_count, 
              candidates_token_count, thoughts_token_count, total_token_count, latency
    
    Returns:
        EfficiencyScores object with aggregated metrics
    """
    if USE_PANDAS:
        # Filter out rows with missing data
        valid_df = data.dropna(subset=['latency', 'prompt_token_count', 'candidates_token_count'])
    else:
        # Convert list of dicts to valid list
        valid_data = [
            row for row in data 
            if row.get('latency') is not None 
            and row.get('prompt_token_count') is not None 
            and row.get('candidates_token_count') is not None
        ]
        valid_df = valid_data
    
    if len(valid_df) == 0:
        print("Warning: No valid rows found in the dataset")
        return EfficiencyScores()
    
    # Calculate total metrics across all requests
    if USE_PANDAS:
        total_latency = valid_df['latency'].sum()
        total_input_tokens = valid_df['prompt_token_count'].sum()
        total_output_tokens = valid_df['candidates_token_count'].sum()
        
        # Add thoughts tokens to output if available
        if 'thoughts_token_count' in valid_df.columns:
            total_thoughts_tokens = valid_df['thoughts_token_count'].fillna(0).sum()
            total_output_tokens += total_thoughts_tokens
    else:
        total_latency = sum(row.get('latency', 0) for row in valid_df)
        total_input_tokens = sum(row.get('prompt_token_count', 0) for row in valid_df)
        total_output_tokens = sum(row.get('candidates_token_count', 0) for row in valid_df)
        
        # Add thoughts tokens to output if available
        total_thoughts_tokens = sum(row.get('thoughts_token_count', 0) or 0 for row in valid_df)
        total_output_tokens += total_thoughts_tokens
    
    # Calculate total cost: input cost + output cost
    input_cost = (total_input_tokens / 1000.0) * INPUT_PRICE_PER_1K
    output_cost = (total_output_tokens / 1000.0) * OUTPUT_PRICE_PER_1K
    total_cost = input_cost + output_cost
    
    # Calculate average latency
    if USE_PANDAS:
        avg_latency = valid_df['latency'].mean()
    else:
        avg_latency = total_latency / len(valid_df) if len(valid_df) > 0 else 0
    
    # Calculate tokens per second (output tokens / total latency)
    if total_latency > 0:
        tokens_per_second = total_output_tokens / total_latency
    else:
        tokens_per_second = 0.0
    
    # Calculate cost per 1k tokens (weighted average)
    total_tokens = total_input_tokens + total_output_tokens
    if total_tokens > 0:
        cost_per_1k = (total_cost / total_tokens) * 1000.0
    else:
        cost_per_1k = 0.0
    
    # Create EfficiencyScores object
    efficiency = EfficiencyScores(
        tokens_per_second=tokens_per_second,
        latency_seconds=avg_latency,
        cost_per_1k_tokens=cost_per_1k,
    )
    
    return efficiency


def print_statistics(data, efficiency: EfficiencyScores):
    """Print detailed statistics about the efficiency metrics."""
    if USE_PANDAS:
        valid_df = data.dropna(subset=['latency', 'prompt_token_count', 'candidates_token_count'])
    else:
        valid_df = [
            row for row in data 
            if row.get('latency') is not None 
            and row.get('prompt_token_count') is not None 
            and row.get('candidates_token_count') is not None
        ]
    
    if len(valid_df) == 0:
        print("No valid data to display statistics")
        return
    
    if USE_PANDAS:
        total_input_tokens = valid_df['prompt_token_count'].sum()
        total_output_tokens = valid_df['candidates_token_count'].sum()
        if 'thoughts_token_count' in valid_df.columns:
            total_thoughts_tokens = valid_df['thoughts_token_count'].fillna(0).sum()
            total_output_tokens += total_thoughts_tokens
        total_tokens = total_input_tokens + total_output_tokens
        total_latency = valid_df['latency'].sum()
        latencies = valid_df['latency'].tolist()
    else:
        total_input_tokens = sum(row.get('prompt_token_count', 0) for row in valid_df)
        total_output_tokens = sum(row.get('candidates_token_count', 0) for row in valid_df)
        total_thoughts_tokens = sum(row.get('thoughts_token_count', 0) or 0 for row in valid_df)
        total_output_tokens += total_thoughts_tokens
        total_tokens = total_input_tokens + total_output_tokens
        total_latency = sum(row.get('latency', 0) for row in valid_df)
        latencies = [row.get('latency', 0) for row in valid_df]
    
    input_cost = (total_input_tokens) * INPUT_PRICE_PER_1K
    output_cost = (total_output_tokens) * OUTPUT_PRICE_PER_1K
    total_cost = input_cost + output_cost
    
    print("=" * 60)
    print("EFFICIENCY STATISTICS")
    print("=" * 60)
    print(f"\nDataset Summary:")
    print(f"  Total requests: {len(valid_df)}")
    print(f"  Valid requests: {len(valid_df)}")
    
    print(f"\nToken Usage:")
    print(f"  Total input tokens: {total_input_tokens:,}")
    print(f"  Total output tokens: {total_output_tokens:,}")
    print(f"  Total tokens: {total_tokens:,}")
    print(f"  Average input tokens per request: {total_input_tokens / len(valid_df):.2f}")
    print(f"  Average output tokens per request: {total_output_tokens / len(valid_df):.2f}")
    
    print(f"\nLatency:")
    print(f"  Total latency: {total_latency:.2f} seconds")
    print(f"  Average latency: {efficiency.latency_seconds:.4f} seconds")
    if USE_PANDAS:
        print(f"  Min latency: {valid_df['latency'].min():.4f} seconds")
        print(f"  Max latency: {valid_df['latency'].max():.4f} seconds")
        print(f"  Median latency: {valid_df['latency'].median():.4f} seconds")
    else:
        print(f"  Min latency: {min(latencies):.4f} seconds")
        print(f"  Max latency: {max(latencies):.4f} seconds")
        sorted_latencies = sorted(latencies)
        median_idx = len(sorted_latencies) // 2
        median = sorted_latencies[median_idx] if len(sorted_latencies) % 2 == 1 else (sorted_latencies[median_idx-1] + sorted_latencies[median_idx]) / 2
        print(f"  Median latency: {median:.4f} seconds")
    
    print(f"\nCost Breakdown:")
    print(f"  Input cost: ${input_cost:.6f} (${INPUT_PRICE_PER_1M}/1M tokens)")
    print(f"  Output cost: ${output_cost:.6f} (${OUTPUT_PRICE_PER_1M}/1M tokens)")
    print(f"  Total cost: ${total_cost:.6f}")
    print(f"  Cost per 1k tokens: ${efficiency.cost_per_1k_tokens:.6f}")
    print(f"  Cost per request: ${total_cost / len(valid_df):.6f}")
    
    print(f"\nEfficiencyScores Object:")
    print(f"  tokens_per_second: {efficiency.tokens_per_second}")
    print(f"  latency_seconds: {efficiency.latency_seconds:.4f}")
    print(f"  cost_per_1k_tokens: ${efficiency.cost_per_1k_tokens:.6f}")
    print("=" * 60)


if __name__ == "__main__":
    # Read the Excel file
    excel_file = Path("gemini/gemini_results1.xlsx")
    
    if not excel_file.exists():
        print(f"Error: File {excel_file} not found")
        exit(1)
    
    print(f"Reading {excel_file}...")
    
    if USE_PANDAS:
        df = pd.read_excel(excel_file)
        print(f"Loaded {len(df)} rows")
        print(f"Columns: {df.columns.tolist()}")
        data = df
    else:
        # Use openpyxl to read Excel
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Read headers
        headers = [cell.value for cell in ws[1]]
        print(f"Columns: {headers}")
        
        # Read data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)
        
        print(f"Loaded {len(data)} rows")
        wb.close()
    
    # Calculate efficiency scores
    efficiency = calculate_gemini_efficiency(data)
    
    # Print statistics
    print_statistics(data, efficiency)
    
    # Return the EfficiencyScores object
    print(f"\nReturned EfficiencyScores object:")
    print(efficiency)

