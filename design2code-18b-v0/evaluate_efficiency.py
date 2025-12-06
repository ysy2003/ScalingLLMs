"""
Evaluate efficiency metrics for Design2Code-18B-V0 local model results from design2code_18b_v0_results1.xlsx
"""
import sys
from pathlib import Path

# Add parent directory to path to import metrics
sys.path.insert(0, str(Path(__file__).parent.parent))

from metrics.efficiency import EfficiencyScores, compute_efficiency_local

try:
    import pandas as pd
    USE_PANDAS = True
except (ImportError, ValueError):
    USE_PANDAS = False

try:
    from openpyxl import load_workbook
    USE_OPENPYXL = True
except ImportError:
    USE_OPENPYXL = False
    if not USE_PANDAS:
        print("Error: Need either pandas or openpyxl to read Excel files")
        exit(1)


def calculate_design2code_efficiency(data) -> EfficiencyScores:
    """
    Calculate efficiency scores for Design2Code-18B-V0 local model results.
    
    Args:
        data: DataFrame or list of dicts with columns: candidates_token_count, latency
    
    Returns:
        EfficiencyScores object with aggregated metrics
    """
    if USE_PANDAS:
        # Filter out rows with missing data
        valid_df = data.dropna(subset=['latency', 'candidates_token_count'])
    else:
        # Convert list of dicts to valid list
        valid_data = [
            row for row in data 
            if row.get('latency') is not None 
            and row.get('candidates_token_count') is not None
        ]
        valid_df = valid_data
    
    if len(valid_df) == 0:
        print("Warning: No valid rows found in the dataset")
        return EfficiencyScores()
    
    # Calculate total metrics across all requests
    if USE_PANDAS:
        total_generated_tokens = valid_df['candidates_token_count'].sum()
        total_wall_time = valid_df['latency'].sum()
        avg_latency = valid_df['latency'].mean()
    else:
        total_generated_tokens = sum(row.get('candidates_token_count', 0) for row in valid_df)
        total_wall_time = sum(row.get('latency', 0) for row in valid_df)
        avg_latency = total_wall_time / len(valid_df) if len(valid_df) > 0 else 0
    
    # Use compute_efficiency_local for local models
    # This calculates tokens_per_second based on total tokens / total time
    efficiency = compute_efficiency_local(
        generated_tokens=total_generated_tokens,
        wall_time_sec=total_wall_time,
        latency_sec=avg_latency,  # Use average latency per request
    )
    
    return efficiency


def print_statistics(data, efficiency: EfficiencyScores):
    """Print detailed statistics about the efficiency metrics."""
    if USE_PANDAS:
        valid_df = data.dropna(subset=['latency', 'candidates_token_count'])
    else:
        valid_df = [
            row for row in data 
            if row.get('latency') is not None 
            and row.get('candidates_token_count') is not None
        ]
    
    if len(valid_df) == 0:
        print("No valid data to display statistics")
        return
    
    if USE_PANDAS:
        total_generated_tokens = valid_df['candidates_token_count'].sum()
        total_wall_time = valid_df['latency'].sum()
        latencies = valid_df['latency'].tolist()
        token_counts = valid_df['candidates_token_count'].tolist()
    else:
        total_generated_tokens = sum(row.get('candidates_token_count', 0) for row in valid_df)
        total_wall_time = sum(row.get('latency', 0) for row in valid_df)
        latencies = [row.get('latency', 0) for row in valid_df]
        token_counts = [row.get('candidates_token_count', 0) for row in valid_df]
    
    print("=" * 60)
    print("EFFICIENCY STATISTICS")
    print("=" * 60)
    print(f"\nDataset Summary:")
    print(f"  Total requests: {len(valid_df)}")
    print(f"  Valid requests: {len(valid_df)}")
    
    print(f"\nToken Usage:")
    print(f"  Total generated tokens: {total_generated_tokens:,}")
    print(f"  Average tokens per request: {total_generated_tokens / len(valid_df):.2f}")
    if USE_PANDAS:
        print(f"  Min tokens: {valid_df['candidates_token_count'].min():,}")
        print(f"  Max tokens: {valid_df['candidates_token_count'].max():,}")
        print(f"  Median tokens: {valid_df['candidates_token_count'].median():.0f}")
    else:
        print(f"  Min tokens: {min(token_counts):,}")
        print(f"  Max tokens: {max(token_counts):,}")
        sorted_tokens = sorted(token_counts)
        median_idx = len(sorted_tokens) // 2
        median = sorted_tokens[median_idx] if len(sorted_tokens) % 2 == 1 else (sorted_tokens[median_idx-1] + sorted_tokens[median_idx]) / 2
        print(f"  Median tokens: {median:.0f}")
    
    print(f"\nLatency:")
    print(f"  Total wall time: {total_wall_time:.2f} seconds")
    print(f"  Average latency: {efficiency.latency_seconds:.4f} seconds")
    if USE_PANDAS:
        print(f"  Min latency: {valid_df['latency'].min():.4f} seconds")
        print(f"  Max latency: {valid_df['latency'].max():.4f} seconds")
        print(f"  Median latency: {valid_df['latency'].median():.4f} seconds")
    else:
        print(f"  Min latency: {min(latencies):.4f} seconds")
        print(f"  Max latency: {max(latencies):.4f} seconds")
        sorted_latencies = sorted(latencies)
        median_idx = len(sorted_latencies) // 2
        median = sorted_latencies[median_idx] if len(sorted_latencies) % 2 == 1 else (sorted_latencies[median_idx-1] + sorted_latencies[median_idx]) / 2
        print(f"  Median latency: {median:.4f} seconds")
    
    print(f"\nPerformance:")
    print(f"  Tokens per second: {efficiency.tokens_per_second:.4f}")
    print(f"  Average tokens per second per request: {total_generated_tokens / total_wall_time:.4f}")
    
    print(f"\nEfficiencyScores Object:")
    print(f"  tokens_per_second: {efficiency.tokens_per_second:.4f}")
    print(f"  latency_seconds: {efficiency.latency_seconds:.4f}")
    print(f"  cost_per_1k_tokens: {efficiency.cost_per_1k_tokens}")
    print("=" * 60)


if __name__ == "__main__":
    # Read the Excel file
    excel_file = Path("design2code-18b-v0/design2code_18b_v0_results1.xlsx")
    
    if not excel_file.exists():
        print(f"Error: File {excel_file} not found")
        exit(1)
    
    print(f"Reading {excel_file}...")
    
    if USE_PANDAS:
        df = pd.read_excel(excel_file)
        print(f"Loaded {len(df)} rows")
        print(f"Columns: {df.columns.tolist()}")
        data = df
    else:
        # Use openpyxl to read Excel
        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Read headers
        headers = [cell.value for cell in ws[1]]
        print(f"Columns: {headers}")
        
        # Read data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)
        
        print(f"Loaded {len(data)} rows")
        wb.close()
    
    # Calculate efficiency scores
    efficiency = calculate_design2code_efficiency(data)
    
    # Print statistics
    print_statistics(data, efficiency)
    
    # Return the EfficiencyScores object
    print(f"\nReturned EfficiencyScores object:")
    print(efficiency)

